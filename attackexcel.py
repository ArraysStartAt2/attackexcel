#!/usr/bin/env python3
"""
Manipulate an Excel workbook for use with the MITRE ATT&CK navigator. The goal here is to make it easy to leverage
every CSOC's favorite tool (Excel) to generate a navigator layer to help visualize control gaps.
"""
import argparse
import json
import sys

from attackcti import attack_client
from openpyxl import load_workbook, Workbook

__author__ = "Lawrence Smith"
__copyright__ = "Copyright 2021, Lawrence Smith"
__license__ = "GPL"
__version__ = "0.0.1"
__maintainer__ = "Lawrence Smith"
__email__ = "63609867+ArraysStartAt2@users.noreply.github.com"
__status__ = "Prototype"

default_layer_template = {
	"versions": {
		"attack": "9",
		"navigator": "4.3",
		"layer": "4.2"
	},
	"domain": "",
	"description": "",
	"filters": {"platforms": []},
	"sorting": 0,
	"layout": {
		"layout": "side",
		"showID": "false",
		"showName": "true"
	},
	"hideDisabled": "false",
	"gradient": {
		"colors": [
			"#ff6666",
			"#ffe766",
			"#8ec843"
		],
		"minValue": 0,
		"maxValue": 100
	},
	"legendItems": [],
	"metadata": [],
	"showTacticRowBackground": "false",
	"tacticRowBackground": "#dddddd",
	"selectTechniquesAcrossTactics": "true",
	"selectSubtechniquesWithParent": "false"
}

valid_enterprise_platforms = {'SaaS', 'macOS', 'PRE', 'IaaS', 'Linux', 'Office 365', 'Containers',
                              'Google Workspace', 'Windows', 'Network', 'Azure AD'}
valid_mobile_platforms = {'Android', 'iOS'}
valid_ics_platforms = {'Field Controller/RTU/PLC/IED', 'Safety Instrumented System/Protection Relay', 'Control Server',
                       'Input/Output Server', 'Windows', 'Human-Machine Interface', 'Engineering Workstation',
                       'Data Historian'}


def validate_platform_filters_to_domain(args):
    """
    Since argparse does not have any ability to ensure that certain combinations of parameters are valid, this
    function does just that.
    """
    platform_filter = None
    if args.platformfilterin is not None:
        platform_filter = args.platformfilterin
    if args.platformfilterout is not None:
        platform_filter = args.platformfilterout
    if platform_filter is None:
        return True
    if args.domain == 'enterprise-attack':
        for platform in platform_filter:
            if platform not in valid_enterprise_platforms:
                print(f'{platform} is not a valid ATT&CK platform for the Enterprise domain')
                return False
    elif args.domain == 'mobile-attack':
        for platform in platform_filter:
            if platform not in valid_mobile_platforms:
                print(f'{platform} is not a valid ATT&CK platform for the Mobile domain')
                return False
    elif args.domain == 'ics-attack':
        for platform in platform_filter:
            if platform not in valid_ics_platforms:
                print(f'{platform} is not a valid ATT&CK platform for the ICS domain')
                return False
    return True


def create_platform_filter(args):
    """
    Creates an appropriate filter depending on the chose domain and filter parameters.
    """
    platform_filter = set()
    if args.domain == 'enterprise-attack':
        platform_filter = valid_enterprise_platforms
        if args.platformfilterin is not None:
            platform_filter = set(args.platformfilterin)
        if args.platformfilterout is not None:
            platform_filter = valid_enterprise_platforms.difference(set(args.platformfilterout))
    elif args.domain == 'mobile-attack':
        platform_filter = valid_mobile_platforms
        if args.platformfilterin is not None:
            platform_filter = set(args.platformfilterin)
        if args.platformfilterout is not None:
            platform_filter = valid_mobile_platforms.difference(set(args.platformfilterout))
    elif args.domain == 'ics-attack':
        platform_filter = valid_ics_platforms
        if args.platformfilterin is not None:
            platform_filter = set(args.platformfilterin)
        if args.platformfilterout is not None:
            platform_filter = valid_ics_platforms.difference(set(args.platformfilterout))

    # We needed to start out with a set to do some set math, but now we want a list
    return platform_filter


def seed(args):
    """
    Downloads the latest ATT&CK framework and loads it into an new Excel workbook specified by 'outfile'. More
    specifically, it will create the following worksheets in a relational format with the following fields:
    techniques (techniqueID, name, description, numberOfDataSources), techniquesToDataSources (techniqueID,
    dataSourceName), dataSources (dataSourceName).
    """

    # Get the ATT&CK techniques of the specified domain
    client = attack_client()
    techniques = dict()
    if args.domain == 'enterprise-attack':
        techniques = client.get_enterprise_techniques(stix_format=False)
    elif args.domain == 'mobile-attack':
        techniques = client.get_mobile_techniques(stix_format=False)
    elif args.domain == 'ics-attack':
        techniques = client.get_ics_techniques(stix_format=False)

    # Create a new workbook and some sheets
    workbook = Workbook()
    sheet1 = workbook.create_sheet(title='techniques')
    sheet2 = workbook.create_sheet(title='techniquesToDataSources')
    sheet3 = workbook.create_sheet(title='dataSources')

    # Create headers
    sheet1.cell(row=1, column=1, value='techniqueID')
    sheet1.cell(row=1, column=2, value='name')
    sheet1.cell(row=1, column=3, value='isSubtechnique')
    sheet1.cell(row=1, column=4, value='platforms')
    sheet1.cell(row=1, column=5, value='description')
    sheet2.cell(row=1, column=1, value='techniqueID')
    sheet2.cell(row=1, column=2, value='dataSourceName')
    sheet3.cell(row=1, column=1, value='dataSourceName')

    # Create a set to hold unique data sources
    data_sources = set()

    # Loop through all the techniques
    sheet1_row = 2
    sheet2_row = 2
    platform_filter = create_platform_filter(args)
    for technique in techniques:
        # skip this technique if it was revoked
        if ('revoked' in technique) and (technique['revoked']):
            print(f'Skipping {technique["technique_id"]} because it was revoked.')
            continue
        # skip this technique if it's a subtechnique and the no-subtechniques flag is set
        if technique['x_mitre_is_subtechnique'] and not args.subtechniques:
            print(f'Skipping {technique["technique_id"]} because it is a subtechnique')
            continue
        # skip this technique if the none of associated platforms are in
        # our platform filter i.e. the two sets are disjoint.
        if platform_filter.isdisjoint(set(technique['platform'])):
            print(f'Skipping {technique["technique_id"]} because none of its platforms {technique["platform"]} '
                  f'match filter.')
            continue

        sheet1.cell(row=sheet1_row, column=1, value=technique['technique_id'])
        sheet1.cell(row=sheet1_row, column=2, value=technique['technique'])
        sheet1.cell(row=sheet1_row, column=3, value=str(technique['x_mitre_is_subtechnique']))
        sheet1.cell(row=sheet1_row, column=4, value=str(technique['platform']))
        if 'data_sources' in technique:
            for data_source in technique['data_sources']:
                sheet2.cell(row=sheet2_row, column=1, value=technique['technique_id'])
                sheet2.cell(row=sheet2_row, column=2, value=data_source)
                data_sources.add(data_source)
                sheet2_row = sheet2_row + 1
        if 'technique_description' in technique:
            sheet1.cell(row=sheet1_row, column=5, value=technique['technique_description'])
        sheet1_row = sheet1_row + 1

    # Loop through the extracted data sources
    sheet3_row = 2
    for data_source in data_sources:
        sheet3.cell(row=sheet3_row, column=1, value=data_source)
        sheet3_row = sheet3_row + 1

    workbook.save(args.outfile)

    print(f'Excel workbook created at \'{args.outfile}\' with {sheet1_row-1} techniques.')


def layer(args):
    """
    Takes a sheet specified by 'sheet' of an Excel document specified by 'infile', looks for the columns
    'techniqueID' (mandatory), 'color', 'enabled', 'score', and 'comment', creates a list of techniques from
    those rows, and creates a json file specified by 'outfile' that is suitable for uploading as a layer in the
    ATT&CK navigator.
    """

    # Load the workbook and the sheet
    workbook = load_workbook(filename=args.infile, read_only=True, data_only=True)
    worksheet = workbook[args.worksheet]

    # Look for the following column headers and create a mapping to their column numbers:
    # techniqueID, color, enabled, score, comment
    column_headers = worksheet[1]
    column_map = dict()
    column = 0
    for cell in column_headers:
        if cell.value in ['techniqueID', 'color', 'enabled', 'score', 'comment']:
            column_map[cell.value] = column
        column = column + 1

    # Create a list of techniques, with each technique structured as a dict where the key is one of the
    # available column headers and the value the value is the value of the cell of the mapped column
    techniques = list()
    for row in worksheet.iter_rows(min_row=2):
        technique = dict()
        for header, column in column_map.items():
            technique[header] = row[column].value
        techniques.append(technique)

    # add the name, description, and list of techniques to our template
    layer_template = default_layer_template
    layer_template['name'] = args.name
    layer_template['domain'] = args.domain
    layer_template['description'] = args.description
    layer_template['techniques'] = techniques
    layer_template['filters']['platforms'] = list(create_platform_filter(args))

    # dump the layer template to a file
    f = open(args.outfile, 'w')
    f.write(json.dumps(layer_template, indent=1))
    f.close()

    print(f'ATT&CK Navigator layer file written to \'{args.outfile}\' with {len(techniques)} techniques')


def main():
    # Create the command-line argument parser
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(title='subcommands',
                                       description='This tool contains the following subcommands, exactly one of '
                                       'which is required.',
                                       help='Use ''attackexcel.py <subcommand> -h'' to get help on each '
                                       'subcommand.',
                                       required=True)

    # create subparser for the 'seed' command
    parser_seed = subparsers.add_parser(name='seed', description=seed.__doc__)
    parser_seed.add_argument('outfile', type=str, help='the path to the Excel workbook.')
    parser_seed.add_argument('--domain', type=str, choices=['enterprise-attack', 'mobile-attack', 'ics-attack'],
                             default='enterprise-attack',
                             help='the ATT&CK domain to download (default=enterprise-attack)')
    parser_seed.add_argument('--no-subtechniques', dest='subtechniques', action='store_false')
    parser_seed.set_defaults(subtechniques=True)
    parser_seed_group_filter = parser_seed.add_mutually_exclusive_group()
    parser_seed_group_filter.add_argument('--platformfilterout', type=str, nargs='+',
                                   choices=list(valid_enterprise_platforms))
    parser_seed_group_filter.add_argument('--platformfilterin', type=str, nargs='+',
                                   choices=set().union(valid_enterprise_platforms,
                                                       valid_mobile_platforms,
                                                       valid_ics_platforms))
    parser_seed.set_defaults(func=seed)

    # create subparser for the 'layer' command
    parser_layer = subparsers.add_parser(name='layer', description=layer.__doc__)
    parser_layer.add_argument('infile', type=str, help='the path to the Excel workbook')
    parser_layer.add_argument('outfile', type=str, help='the path to the json document to be created')
    parser_layer.add_argument('--worksheet', type=str, default='techniques',
                              help='the name of the worksheet within the identified workbook')
    parser_layer.add_argument('--domain', type=str, choices=['enterprise-attack', 'mobile-attack', 'ics-attack'],
                              default='enterprise-attack',
                              help='the ATT&CK domain of the layer to create (default=\'enterprise-attack\')')
    parser_layer.add_argument('--name', type=str, default='Attackexcel Layer',
                              help='name of the layer (default=\'Attackexcel Layer\')')
    parser_layer.add_argument('--description', type=str, default='',
                              help='a description for the layer (default=none)')
    parser_layer_group = parser_layer.add_mutually_exclusive_group()
    parser_layer_group.add_argument('--platformfilterout', type=str, nargs='+',
                                    choices=list(valid_enterprise_platforms))
    parser_layer_group.add_argument('--platformfilterin', type=str, nargs='+',
                                    choices=set().union(valid_enterprise_platforms,
                                                        valid_mobile_platforms,
                                                        valid_ics_platforms))
    parser_layer.set_defaults(func=layer)

    # parse the arguments
    if len(sys.argv) == 1:
        parser.print_help(sys.stderr)
        sys.exit(1)
    args = parser.parse_args()

    # validate domain to platform match, and call the function to handle the appropriate command
    if validate_platform_filters_to_domain(args):
        args.func(args)


if __name__ == '__main__':
    main()
